import os
import sys
import six
from datetime import datetime
import ExtDataSources
import xlrd
import traceback
from xlrd import open_workbook
from xlwt import Workbook
from xlutils.copy import copy
import yfinance as yf
from polygon import RESTClient
from GetStockRank import GetStockRank
from GetTargetStop import GetTargetStop
import openpyxl
import json
#import DBMethods

MINUTES_BETWEEN_HOD_TRIGGER = 45
REQUIRED_BO_LEVEL = .15


class EODAnalysis(object):
    def __init__(self):
        self.gsr = GetStockRank()
        self.gts = GetTargetStop()
        self.client = RESTClient()

    def GetSavedData(self,Ticker):
        data = {}
        TextFile = "%s%s.txt"%(os.getenv('AFTERNOONBODATA'),Ticker)
        if os.path.exists(TextFile): #Read Date/time
            with open(TextFile, "r") as rd:
                for line in rd:
                    line = line.replace("u'","'")
                    line = line.replace("'","\"")
                    key = line.split(" {")[0]
                    payload = json.loads(line.split("00 ")[1])
                    data[key] = payload
        return data
                    
                    
    def SaveAlphaDataToAccess(self,Ticker,data):
        #extData = ExtDataSources.ExtDataSources()
        #dbmethods = DBMethods.DBMethods(os.getenv('AFTERNOONBOHISTORYDB'))
        #DATE_LOCATION = 0
        #TICKER_LOCATION = 1
        #if os.path.exists(os.getenv('AFTERNOONBOTRACKING')):
        #    rb = open_workbook(os.getenv('AFTERNOONBOTRACKING'))
        #    rs = rb.sheet_by_name("Tracking")
        #    for row_idx in range(1, 2):
        #    #for row_idx in range(27, 28):
        #        Ticker = rs.cell(row_idx,TICKER_LOCATION).value
        #        tDate = rs.cell(row_idx,DATE_LOCATION).value
        #        datetrigger = "%s 09:31:00"%tDate
        #        data,metadata = extData.getLiveTickerIntraday_values(Ticker, "1min", "full", "json", timeout=10, trigger=datetrigger)
                #TextFile = r'C:\Users\Kevin\Documents\StockData\AfternoonBO\%s.txt'%Ticker
                TextFile = "%s%s.txt"%(os.getenv('AFTERNOONBODATA'),Ticker)
                #print(TextFile)
                #Get Values of interest#
                row = 0 
                LastEntryDateTime = None
                ExistingFile = False
                if data:
                    if os.path.exists(TextFile): #Read Date/time 
                        ExistingFile = True
                        with open(TextFile, "r") as rd:
                            line = None
                            for line in rd:
                                pass
                            last_line = line
                            #Date Format 2021-11-10 08:18:00 #
                            sDateTime = "%s %s"%(last_line.split(" ")[0],last_line.split(" ")[1])
                            print(sDateTime)
                            LastEntryDateTime = datetime.strptime(sDateTime,"%Y-%m-%d %H:%M:%S") #Only save 
                    with open(TextFile, 'a+') as f:
                        #for key,value in sorted(six.iteritems(data)): #python2.7
                        for key,value in sorted(six.iteritems(data)): #python3.4
                            lineDateTime = datetime.strptime(key,"%Y-%m-%d %H:%M:%S")
                            if not ExistingFile or (lineDateTime > LastEntryDateTime):
                                try:
                                    linetowrite = "%s %s\n"%(key,value)
                                    print(linetowrite)
                                    f.write(linetowrite)
                                    #if not dbmethods.TableExists(Ticker):
                                    #    #CreateTable#
                                    #    print("Trying to create table")
                                    #    #dbmethods.CreateTable(Ticker)
                                    #dbmethods.insertRowIntoTable("%s"%Ticker,
                                    #                    key,
                                    #                    value["1. open"],
                                    #                    value["2. high"],
                                    #                    value["3. low"],
                                    #                    value["4. close"],
                                    #                    value["5. volume"],
                                    #                    metadata["4. Interval"]
                                    #                    )
                                except Exception as detail:
                                    print(detail)
                                    continue
                                    #print "Table %s had an insertion error; probably duplicate row entry"%SELECT_STOCK
                                    #print detail
                
    def FindMostRecentKey(self, data, day, h, m, stop):
        time = "%d:%d:00"%(h,m)
        key = "%s %s"%(day,time)
        while key not in data.keys() and h >= stop:  #Dataset sometimes missing minutes; Count backwards from 16:00 until 12:00 looking for the most recent value available for estimated prevClose
            time = "%02d:%02d:00"%(h,m)
            if m == 0:
                h = h - 1
                m = 59
            else:
                m = m - 1
            key = "%s %s"%(day,time)
            print(key)
        return key
        
    def ProcessAlphVantageData(self,ticker,BOdate):
        extData = ExtDataSources.ExtDataSources()
        datetrigger = "%s 09:31:00"%BOdate
        data = self.GetSavedData(ticker)
        if datetrigger not in data:
            data,metadata = extData.getLiveTickerIntraday_values(ticker, "1min", "full", "json", timeout=120, trigger=datetrigger)
            self.SaveAlphaDataToAccess(ticker,data)
        #Get Values of interest#
        if data:
            day_volume = 0
            dayHigh = 0
            tHOD = 0
            tHOD_Break = 0
            tHOD_Initial = 0
            tDrawdown_Max = 0
            Breakout_Level = 0
            Breakout_Volume = 0
            Drawdown = 0
            Close = 0
            premarket = True
            marketOpen = True
            ListOfDates = []
            PreviousClose = 0
            dayHighGain = 0
            Closekey = None
            for key,value in sorted(six.iteritems(data)):
                #print(key,value)
                day = key.split(" ")[0]
                if day not in ListOfDates:
                    ListOfDates.append(key.split(" ")[0])
                if day == BOdate and marketOpen:
                    #Calculate all critical values (Initial HOD/Initial tHOD/Volume @ HOD Break/BO_Time/HOD tHOD/Close/MaxDrawdown)
                    if PreviousClose == 0:
                        PreviousDay = ListOfDates[-2]
                        PreviousCloseKey = self.FindMostRecentKey(data,PreviousDay,16,00,12)
                        try:
                            PreviousClose = float(data[PreviousCloseKey]["4. close"])
                            print("Previous Close: %s"%PreviousClose)
                        except Exception as detail:
                            print("Exception thrown getting Previous Close")
                            print("Key: %s"%key)
                            print(detail)
                    day_volume += int(value["5. volume"])
                    datetime_object = datetime.strptime(key,"%Y-%m-%d %H:%M:%S")
                    if premarket and ("09:30:00" in key) or ("09:31:00" in key):  #This is because alpha_vantage morning minutes are not reliable. Hoping for one of these minutes to be present.
                        premarket = False
                        tHOD_Initial = datetime_object
                        dayHigh = float(value["2. high"])
                        dayHighGain = float(dayHigh - PreviousClose)/PreviousClose
                    elif not premarket:
                        if float(value["2. high"]) > dayHigh:
                            if (tHOD_Break == 0) and (divmod((datetime_object - tHOD_Initial).total_seconds(), 60)[0] >= MINUTES_BETWEEN_HOD_TRIGGER and dayHighGain >= REQUIRED_BO_LEVEL):
                                tHOD_Break = datetime_object
                                Breakout_Level = dayHigh
                                Drawdown = Breakout_Level
                                Breakout_Volume = day_volume
                            elif tHOD_Break == 0:
                                tHOD_Initial = datetime_object
                            else:
                                tHOD = datetime_object
                            dayHigh = float(value["2. high"])
                            dayHighGain = float(dayHigh - PreviousClose)/PreviousClose
                        if Breakout_Level is not None:
                            if float(value["3. low"]) < Drawdown:
                                Drawdown = float(value["3. low"])
                                tDrawdown_Max = datetime_object
                        if Close == 0:
                            Closekey = self.FindMostRecentKey(data,BOdate,16,00,12)  #Find last valid time entry in data set and this is our close key
                            Close = float(data[Closekey]["4. close"])
                    
                if Closekey == key:           #NOT ACCOUNTING FOR 1/2 DAYS THIS WILL TAKE AFTERMARKET DATA ON HALF DAYS#
                    marketOpen = False
            #if Close == 0:
            #    try:
            #        Close = float(data["%s %s"%(BOdate,"13:00:00")]["4. close"])  #Get Half-day closing value; Need to clean up this implementation of half-days!#
            #    except:
            #        Close = float(data["%s %s"%(BOdate,"13:01:00")]["4. close"])  #Get Half-day closing value; Need to clean up this implementation of half-days!#
            print(tHOD_Initial,tHOD_Break,Breakout_Level,Breakout_Volume,tHOD,dayHigh,tDrawdown_Max,Drawdown,Close,PreviousClose)
            return tHOD_Initial,tHOD_Break,Breakout_Level,Breakout_Volume,tHOD,dayHigh,tDrawdown_Max,Drawdown,Close,PreviousClose
        
    def ProcessDailyBreakouts(self):
        DATE_LOCATION = 0
        TICKER_LOCATION = 1
        STKFLOAT = 2
        MARKETCAP = 4
        BO_LEVEL = 11
        HOD_LOCATION = 12
        PREV_CLOSE_LOCATION = 17
        BO_GAIN = 21
        RANK = 22
        PROFIT = 23
        DATA_SAVED = 24
        
        if os.path.exists(os.environ["AFTERNOONBOTRACKING"]):
            rb = open_workbook(os.environ["AFTERNOONBOTRACKING"])
            rs = rb.sheet_by_name("Tracking")
            wb = copy(rb)
            ws = wb.get_sheet("Tracking")
            for row_idx in range(1, rs.nrows):
            #for row_idx in range(1, 23):
                Ticker = rs.cell(row_idx,TICKER_LOCATION).value
                Hod = rs.cell(row_idx,HOD_LOCATION).value
                tDate = rs.cell(row_idx,DATE_LOCATION).value
                PrevClose = rs.cell(row_idx,PREV_CLOSE_LOCATION).value
                BO_Gap = rs.cell(row_idx,BO_GAIN).value
                rank = rs.cell(row_idx,RANK).value
                Breakout_Level = rs.cell(row_idx,BO_LEVEL).value
                
                if BO_Gap == "" and Breakout_Level != "" and PrevClose != "":
                    BO_Gap = (Breakout_Level - PrevClose)/PrevClose
                    ws.write(row_idx,BO_GAIN,BO_Gap)
                    wb.save(os.environ["AFTERNOONBOTRACKING"])
                elif BO_Gap == "":  #Need a float so we can compare to see if the BO_Gap is big enough.
                    BO_Gap = 0.0

                #y, m, d, h, i, s = datetime.datetime(*xlrd.xldate_as_tuple(tDate, rb.datemode))
                if Hod == "" or float(BO_Gap) < REQUIRED_BO_LEVEL:
                    #Request Alpha_vantage Data
                    print(Ticker, Hod, tDate)
                    #if "floatShares" in YahooInfo.info and YahooInfo.info["floatShares"] is not None:
                    try:
                        #YahooInfo = yf.Ticker(Ticker)
                        #Sharefloat = int(YahooInfo.info["floatShares"])
                        PolygonInfo = self.client.get_ticker_details(Ticker)
                        Sharefloat = PolygonInfo.weighted_shares_outstanding
                        tHOD_Initial,tHOD_Break,Breakout_Level,Breakout_Volume,tHOD,dayHigh,tDrawdown_Max,Drawdown,Close,PrevClose = self.ProcessAlphVantageData(Ticker,tDate)
                        if tHOD_Initial: #If Valid Data back From ALPHA
                            #ws.write(row_idx,0,"Date")
                            #ws.write(row_idx,1,"Stock")
                            ws.write(row_idx,2,Sharefloat)
                            ws.write(row_idx,3,str(int(Sharefloat/1000000))+"M")
                            ws.write(row_idx,4,Sharefloat*Breakout_Level) #MarketCap @ BO Level
                            ws.write(row_idx,5,str(int(Sharefloat*Breakout_Level/1000000))+"M") #MarketCap @ BO Level
                            ws.write(row_idx,6,Breakout_Volume)
                            if Sharefloat>0 and Breakout_Level>0:
                                ws.write(row_idx,7,Breakout_Volume*1000000000/(Sharefloat*Breakout_Level))  # Volume/$B Market Cap
                                ws.write(row_idx,8,str(int(Breakout_Volume*1000/(Sharefloat*Breakout_Level)))+"M")  # Volume/$B Market Cap Key
                            ws.write(row_idx,9,tHOD_Initial.strftime("%H:%M:%S"))
                            if tHOD_Break == 0:
                                ws.write(row_idx,10,"NA")
                            else:
                                ws.write(row_idx,10,tHOD_Break.strftime("%H:%M:%S"))
                            ws.write(row_idx,11,Breakout_Level)
                            ws.write(row_idx,12,dayHigh)
                            ws.write(row_idx,13,Close)
                            if Breakout_Level > 0:
                                ws.write(row_idx,14,(dayHigh-Breakout_Level)/Breakout_Level)
                                ws.write(row_idx,15,(Close-Breakout_Level)/Breakout_Level)
                                ws.write(row_idx,16,(Drawdown-Breakout_Level)/Breakout_Level)
                                BO_Gap = (Breakout_Level - PrevClose)/PrevClose
                                ws.write(row_idx,BO_GAIN,BO_Gap)
                            else:
                                ws.write(row_idx,14,"0")
                                ws.write(row_idx,15,"0")
                                ws.write(row_idx,16,"0")
                                ws.write(row_idx,BO_GAIN,"0")
                            ws.write(row_idx,17,PrevClose)
                            #ws.write(row_idx,18,"16:00:00")
                            rank = int(self.gsr.getStockRank(int(Sharefloat),Sharefloat*Breakout_Level)) #float and marketcap
                            ws.write(row_idx,RANK,rank)
                            ws.write(row_idx,24,"Y")
                            #ws.write(row_idx,19,"Options")
                            #ws.write(row_idx,20,"Strike")
                            wb.save(os.environ["AFTERNOONBOTRACKING"])
                    except Exception as detail:
                        print(detail)
                        print(traceback.print_exc(file=sys.stdout))
                        print("Yahoo Finance Missing Critical Data To Analyze, Stock ignored")
                        
                elif PrevClose == "":
                    #ADD CLOSING TIME FOR EXISTING ENTRIES#
                    #ADD PREVIOUS CLOSE#
                    print(Ticker, Hod, tDate)
                    tHOD_Initial,tHOD_Break,Breakout_Level,Breakout_Volume,tHOD,dayHigh,tDrawdown_Max,Drawdown,Close,PreviousClose = self.ProcessAlphVantageData(Ticker,tDate)
                    ws.write(row_idx,17,PreviousClose)
                    ws.write(row_idx,18,"16:00:00")
                    wb.save(os.environ["AFTERNOONBOTRACKING"])
                elif rank == "":
                    stkfloat = int(rs.cell(row_idx,STKFLOAT).value)
                    marketCap = int(rs.cell(row_idx,MARKETCAP).value)
                    rank = int(self.gsr.getStockRank(stkfloat,marketCap))
                    ws.write(row_idx,RANK,rank)
                    wb.save(os.environ["AFTERNOONBOTRACKING"])
                    
    def CalculateProfits(self):
        DATE_LOCATION = 0
        TICKER_LOCATION = 1
        STKFLOAT = 2
        MARKETCAP = 4
        BO_LEVEL = 11
        HOD_LOCATION = 12
        CLOSE = 13
        MAXPR = 14
        CLOSEPR = 15
        DRAWDOWN = 16
        PREV_CLOSE_LOCATION = 17
        BO_GAIN = 21
        RANK = 22
        PROFIT = 23
        DATA_SAVED = 24
        SumOfProfits = 0.0
        
        
        if os.path.exists("%sx"%os.environ["AFTERNOONBOTRACKING"]):
            wb = openpyxl.load_workbook("%sx"%os.environ["AFTERNOONBOTRACKING"])
            ws = wb['Analysis']
            for row in ws.iter_rows(min_row=2):
                rank = row[RANK].value
                stkfloat = 0
                if row[STKFLOAT].value:                    #Checking to make sure we have an entry for this
                    stkfloat = float(row[STKFLOAT].value)
                bo_percent = 0
                if row[BO_GAIN].value:
                    bo_percent = float(row[BO_GAIN].value)
                maxprofit = row[MAXPR].value
                closeprofit = row[CLOSEPR].value
                drawdown = row[DRAWDOWN].value
                calcprofit = row[PROFIT].value
                if stkfloat > 1000000 and bo_percent < 1.00:  #Minimum criteria#
                    target, stop, holidaytarget, holidaystop, marketmin = self.gts.getTargetStop(rank)
                    if drawdown <= stop:
                        profit = stop            #Drawdown limit hit; take loss
                    elif target == "EOD":
                        profit = closeprofit     #Hold trade till end of day
                    elif maxprofit > target:
                        profit = target          #Hit target goal
                    else:
                        profit = closeprofit     #Didn't hit stop or target
                else:
                    profit = 0                   #Don't trade these corner cases
                row[PROFIT].value = profit
                print("Profit: %.2f"%profit)
                SumOfProfits = SumOfProfits + profit
            wb.save("%sx"%os.environ["AFTERNOONBOTRACKING"])
            print("Sum of profits: %.2f"%SumOfProfits)

if __name__ == "__main__":
    import argparse
    analysis = EODAnalysis()
    parser = argparse.ArgumentParser()
    parser.add_argument("-profits", "--profits", help="Run manually to analyze profits", action="store_true")
    args = parser.parse_args()

    if args.profits:
        analysis.CalculateProfits()
    else:
        analysis.ProcessDailyBreakouts()
    #analysis.SaveAlphaDataToAccess()